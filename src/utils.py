"""
Utility functions for Voice to Tag pipeline.
Excel export, data processing, and helper functions.
"""

import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def load_csv_data(
    filepath: str,
    encoding: str = 'utf-8'
) -> pd.DataFrame:
    """
    Load transcription data from CSV file.
    
    Args:
        filepath: Path to CSV file
        encoding: File encoding
        
    Returns:
        DataFrame with transcription data
    """
    df = pd.read_csv(filepath, encoding=encoding)
    
    # Standardize column names
    df.columns = [col.strip() for col in df.columns]
    
    return df


def results_to_dataframe(
    original_df: pd.DataFrame,
    results: List[Dict],
    id_col: str = "ID"
) -> pd.DataFrame:
    """
    Merge extraction results with original data.
    
    Args:
        original_df: Original transcription DataFrame
        results: List of extraction result dicts
        id_col: Column name for client ID
        
    Returns:
        Merged DataFrame with all data
    """
    # Create results DataFrame
    results_df = pd.DataFrame(results)
    
    # Rename client_id to match original
    if 'client_id' in results_df.columns:
        results_df = results_df.rename(columns={'client_id': id_col})
    
    # Convert list columns to strings for Excel
    list_columns = ['tags', 'invalid_tags', 'key_dates', 'dietary', 'allergies', 'mentioned_persons']
    for col in list_columns:
        if col in results_df.columns:
            results_df[col] = results_df[col].apply(
                lambda x: ', '.join(x) if isinstance(x, list) and all(isinstance(i, str) for i in x)
                else json.dumps(x, ensure_ascii=False) if isinstance(x, list) else x
            )
    
    # Merge with original data
    merged_df = original_df.merge(results_df, on=id_col, how='left')
    
    return merged_df


def export_to_excel(
    df: pd.DataFrame,
    output_path: str,
    sheet_name: str = "Tagged Data"
) -> None:
    """
    Export DataFrame to formatted Excel file.
    
    Args:
        df: DataFrame to export
        output_path: Output Excel file path
        sheet_name: Name of the sheet
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            
            if r_idx == 1:
                # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            else:
                cell.alignment = cell_alignment
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, min(len(str(cell.value)), 50))
            except:
                pass
        
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save
    wb.save(output_path)
    print(f"✅ Exported to: {output_path}")


def export_stats_json(
    stats: Dict,
    output_path: str
) -> None:
    """Export extraction statistics to JSON file."""
    stats['exported_at'] = datetime.now().isoformat()
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)
    
    print(f"✅ Stats exported to: {output_path}")


def print_extraction_summary(stats: Dict) -> None:
    """Print a formatted summary of extraction results."""
    print("\n" + "="*60)
    print("📊 EXTRACTION SUMMARY")
    print("="*60)
    
    print(f"\n📝 Processed: {stats.get('total_processed', 0)} notes")
    print(f"🏷️  Total tags extracted: {stats.get('total_tags_extracted', 0)}")
    print(f"🎯 Unique tags used: {stats.get('unique_tags_used', 0)}")
    print(f"📈 Avg tags per note: {stats.get('avg_tags_per_note', 0):.1f}")
    print(f"✨ Avg confidence: {stats.get('avg_confidence', 0):.1%}")
    print(f"💾 From cache: {stats.get('from_cache', 0)}")
    
    print("\n🔝 TOP 10 TAGS:")
    for tag, count in stats.get('top_10_tags', []):
        print(f"   {tag}: {count}")
    
    print("\n📁 TAGS BY CATEGORY:")
    for category, count in sorted(stats.get('tags_by_category', {}).items(), key=lambda x: x[1], reverse=True):
        print(f"   {category}: {count}")
    
    print("\n" + "="*60)


def clear_cache(cache_dir: str = "cache") -> int:
    """
    Clear all cached extraction results.
    
    Args:
        cache_dir: Path to cache directory
        
    Returns:
        Number of files deleted
    """
    cache_path = Path(cache_dir)
    if not cache_path.exists():
        return 0
    
    count = 0
    for file in cache_path.glob("*.json"):
        file.unlink()
        count += 1
    
    print(f"🗑️  Cleared {count} cached files")
    return count


def detect_language(text: str) -> str:
    """
    Simple heuristic language detection based on common words.
    
    Args:
        text: Text to analyze
        
    Returns:
        Language code (FR, EN, IT, ES, DE)
    """
    text_lower = text.lower()
    
    # Common words by language
    indicators = {
        'FR': ['le ', 'la ', 'les ', 'de ', 'du ', 'des ', 'et ', 'est ', 'pour ', 'avec ', 'elle ', 'lui '],
        'EN': ['the ', 'and ', 'is ', 'for ', 'with ', 'she ', 'he ', 'they ', 'her ', 'his '],
        'IT': [' il ', ' la ', ' le ', ' di ', ' del ', ' che ', ' per ', ' con ', ' una ', ' sono '],
        'ES': [' el ', ' la ', ' los ', ' de ', ' del ', ' que ', ' para ', ' con ', ' una ', ' muy '],
        'DE': [' der ', ' die ', ' das ', ' und ', ' ist ', ' für ', ' mit ', ' sie ', ' ihr ', ' sehr ']
    }
    
    scores = {}
    for lang, words in indicators.items():
        scores[lang] = sum(1 for word in words if word in text_lower)
    
    if max(scores.values()) == 0:
        return 'EN'  # Default to English
    
    return max(scores, key=scores.get)


def format_tags_for_display(tags: List[str], max_display: int = 5) -> str:
    """Format tags list for console display."""
    if not tags:
        return "(none)"
    
    if len(tags) <= max_display:
        return ", ".join(tags)
    
    return ", ".join(tags[:max_display]) + f" (+{len(tags) - max_display} more)"
